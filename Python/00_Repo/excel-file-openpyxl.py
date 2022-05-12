# Working with excel files using openpyxl package

# Reviewing an excel file that has multiple worksheets
# Create connection to AWS Athena
# Import utilities, PyAthena and other modules
import pandas as pd
import numpy as np
import matplotlib.pyplot
import seaborn as sns
import time
import sys

# Installing the package needed to connect with the excel file
%%capture
!pip install openpyxl

# Reference website: https://www.freecodecamp.org/news/how-to-create-read-update-and-search-through-excel-files-using-python-c70680d811d4/
import openpyxl

# Load the workbook
excelFile = 'FileReference.xlsx'
excel = openpyxl.load_workbook(excelFile)
print(excel.sheetnames)

# change selection of sheet
activeSheet = excel['Sheet1']
print(activeSheet['B1'].value)
print(activeSheet['C596'].value)

# Review the same cell reference across each worksheet
sheetNames = [x for x in excel.sheetnames]
for i, x in enumerate(sheetNames):
    print('#: {}, current sheet: {}, cell reference A1: {}, max row: {}'.format(i,x,excel[x]['A1'].value,excel[x].max_row))
# max_row : relates to the final cell row value when using the excel command (ctrl + end).
#           This command highlights the final active cell in the worksheet. Which can be larger than the
#           data contained within the worksheet
# Alternative method is to use the pandas method read_excel to import the excel tables into DataFrames
# Examples to follow below on the different imports

# Use read_excel from pandas
df = pd.read_excel(excelFile, sheet_name='Sheet1')
print(df.shape)
df.head()

# Review only the filtered data required
df1 = df[(df['VARIABLE'].str.contains('FILTER',na=False) & (df['STATUS']!='R'))]
df1

# Issue is with the first row, this can be skipped
df2 = pd.read_excel(excelFile, sheet_name='Sheet2', skiprows=1)
df2
